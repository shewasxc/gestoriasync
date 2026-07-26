"""Builds the final Informe_Trimestral_Modelo.xlsx.

Everything downstream of the raw movement values is a *live* Excel formula
(SUMIFS for the quarterly/category pivot, VLOOKUP for the NIF search tool,
SUM/COUNTIF for totals) rather than pre-computed numbers, so a Gestoría can
open the file, tweak a row, and watch the summary recalculate the way an
accountant expects.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models import ExpenseCategory, Movement, ParseIssue

# --- Fintech dark palette, reused from the Streamlit theme -----------------
NAVY = "0B1F3A"
GRAPHITE = "1E2530"
ACCENT_BLUE = "2E86FF"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F4F8"
RED_BG = "FFC7CE"
RED_TEXT = "9C0006"
YELLOW_BG = "FFEB9C"
YELLOW_TEXT = "9C6500"
GREEN_BG = "C6EFCE"
GREEN_TEXT = "006100"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=NAVY)
SUBTOTAL_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)
SUBTOTAL_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
THIN = Side(style="thin", color="B8C0CC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY_FMT = '#,##0.00" €"'
DATE_FMT = "dd/mm/yyyy"

MOV_HEADERS = [
    "Fecha", "NIF/CIF", "NIF Valido", "Concepto", "Importe",
    "Categoria", "Trimestre", "Banco", "Archivo Origen", "Fila Origen",
]
MOV_FIRST_DATA_ROW = 2
MOV_MAX_ROW = 5000  # generous fixed range so formulas keep working after manual edits


def _autosize(ws: Worksheet, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _style_header_row(ws: Worksheet, row: int, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _write_movements_sheet(wb: Workbook, movements: list[Movement]) -> Worksheet:
    ws = wb.active
    ws.title = "Movimientos"

    ws.merge_cells("A1:J1")
    ws["A1"] = "Informe Trimestral — Movimientos Normalizados"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    header_row = 2
    for col_idx, header in enumerate(MOV_HEADERS, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _style_header_row(ws, header_row, len(MOV_HEADERS))

    start = header_row + 1
    for i, m in enumerate(movements):
        r = start + i
        ws.cell(row=r, column=1, value=m.fecha).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=m.nif_cif or "")
        ws.cell(row=r, column=3, value=m.nif_cif_valido if m.nif_cif else None)
        ws.cell(row=r, column=4, value=m.concepto)
        ws.cell(row=r, column=5, value=m.importe).number_format = CURRENCY_FMT
        ws.cell(row=r, column=6, value=m.categoria.value)
        # Live formula: quarter derived from the date, recalculates on edit.
        ws.cell(row=r, column=7, value=f"=ROUNDUP(MONTH(A{r})/3,0)")
        ws.cell(row=r, column=8, value=m.banco.value)
        ws.cell(row=r, column=9, value=m.archivo_origen)
        ws.cell(row=r, column=10, value=m.fila_origen)
        for col_idx in range(1, len(MOV_HEADERS) + 1):
            ws.cell(row=r, column=col_idx).border = BORDER

    last_row = max(start, start + len(movements) - 1)

    # Totals row directly under the data, using a real SUM formula.
    total_row = last_row + 2
    ws.cell(row=total_row, column=4, value="TOTAL").font = SUBTOTAL_FONT
    ws.cell(row=total_row, column=5, value=f"=SUM(E{start}:E{last_row})").number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=5).font = SUBTOTAL_FONT
    ws.cell(row=total_row, column=4).fill = SUBTOTAL_FILL
    ws.cell(row=total_row, column=5).fill = SUBTOTAL_FILL

    # Conditional formatting: red row-tint wherever NIF/CIF failed validation.
    nif_range = f"A{start}:J{MOV_MAX_ROW}"
    ws.conditional_formatting.add(
        nif_range,
        FormulaRule(
            formula=[f"$C{start}=FALSE"],
            fill=PatternFill("solid", fgColor=RED_BG),
            font=Font(color=RED_TEXT),
        ),
    )
    ws.conditional_formatting.add(
        f"E{start}:E{MOV_MAX_ROW}",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color=RED_TEXT)),
    )

    ws.freeze_panes = f"A{start}"
    ws.auto_filter.ref = f"A{header_row}:J{last_row}"
    _autosize(ws, {
        "A": 12, "B": 14, "C": 11, "D": 42, "E": 14, "F": 22,
        "G": 11, "H": 14, "I": 24, "J": 12,
    })
    return ws


def _write_resumen_sheet(wb: Workbook, n_movements: int) -> Worksheet:
    ws = wb.create_sheet("Resumen")
    ws.merge_cells("A1:F1")
    ws["A1"] = "Resumen por Trimestre y Categoria"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    header_row = 3
    headers = ["Categoria", "Trimestre 1", "Trimestre 2", "Trimestre 3", "Trimestre 4", "Total"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _style_header_row(ws, header_row, len(headers))

    mov_start = MOV_FIRST_DATA_ROW
    mov_end = MOV_MAX_ROW
    amount_range = f"Movimientos!$E${mov_start}:$E${mov_end}"
    category_range = f"Movimientos!$F${mov_start}:$F${mov_end}"
    quarter_range = f"Movimientos!$G${mov_start}:$G${mov_end}"

    categories = [c.value for c in ExpenseCategory]
    start = header_row + 1
    for i, category in enumerate(categories):
        r = start + i
        ws.cell(row=r, column=1, value=category)
        for q in range(1, 5):
            col = 1 + q
            formula = f'=SUMIFS({amount_range},{category_range},$A{r},{quarter_range},{q})'
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = CURRENCY_FMT
        total_cell = ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        total_cell.number_format = CURRENCY_FMT
        total_cell.font = SUBTOTAL_FONT
        for col_idx in range(1, 7):
            ws.cell(row=r, column=col_idx).border = BORDER

    last_row = start + len(categories) - 1
    grand_row = last_row + 1
    ws.cell(row=grand_row, column=1, value="TOTAL GENERAL").font = SUBTOTAL_FONT
    for col_letter in ("B", "C", "D", "E", "F"):
        cell = ws.cell(row=grand_row, column="BCDEF".index(col_letter) + 2,
                        value=f"=SUM({col_letter}{start}:{col_letter}{last_row})")
        cell.number_format = CURRENCY_FMT
        cell.font = SUBTOTAL_FONT
        cell.fill = SUBTOTAL_FILL
    ws.cell(row=grand_row, column=1).fill = SUBTOTAL_FILL

    # Quick stats block, also live formulas against the source sheet.
    stats_row = grand_row + 3
    ws.cell(row=stats_row, column=1, value="Movimientos totales").font = SUBTOTAL_FONT
    ws.cell(row=stats_row, column=2, value=f'=COUNTA(Movimientos!D{mov_start}:D{mov_end})')
    ws.cell(row=stats_row + 1, column=1, value="NIF/CIF invalidos").font = SUBTOTAL_FONT
    ws.cell(row=stats_row + 1, column=2,
            value=f'=COUNTIF(Movimientos!C{mov_start}:C{mov_end},FALSE)')
    ws.cell(row=stats_row + 1, column=2).font = Font(color=RED_TEXT, bold=True)

    _autosize(ws, {"A": 26, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
    return ws


def _write_consulta_sheet(wb: Workbook) -> Worksheet:
    """A small live NIF/CIF lookup tool built entirely on VLOOKUP."""
    ws = wb.create_sheet("Consulta NIF")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Buscador de Movimientos por NIF/CIF"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    ws["A3"] = "Introduce un NIF/CIF:"
    ws["A3"].font = SUBTOTAL_FONT
    ws["B3"].fill = PatternFill("solid", fgColor="FFF3CD")
    ws["B3"].border = BORDER
    ws["B3"].alignment = Alignment(horizontal="center")

    mov_range = f"Movimientos!$B${MOV_FIRST_DATA_ROW}:$J${MOV_MAX_ROW}"
    labels_formulas = [
        ("Concepto (primera coincidencia)", f'=IFERROR(VLOOKUP($B$3,{mov_range},3,FALSE),"No encontrado")'),
        ("Importe (primera coincidencia)", f'=IFERROR(VLOOKUP($B$3,{mov_range},4,FALSE),"No encontrado")'),
        ("Categoria", f'=IFERROR(VLOOKUP($B$3,{mov_range},5,FALSE),"No encontrado")'),
        ("Suma total de importes para este NIF/CIF",
         f'=SUMIF(Movimientos!$B${MOV_FIRST_DATA_ROW}:$B${MOV_MAX_ROW},$B$3,'
         f'Movimientos!$E${MOV_FIRST_DATA_ROW}:$E${MOV_MAX_ROW})'),
        ("Numero de movimientos",
         f'=COUNTIF(Movimientos!$B${MOV_FIRST_DATA_ROW}:$B${MOV_MAX_ROW},$B$3)'),
    ]
    for i, (label, formula) in enumerate(labels_formulas):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=r, column=2, value=formula)
        cell.border = BORDER
        if "Importe" in label or "Suma" in label:
            cell.number_format = CURRENCY_FMT

    _autosize(ws, {"A": 36, "B": 22})
    return ws


def _write_validacion_sheet(wb: Workbook, issues: list[ParseIssue]) -> Worksheet:
    ws = wb.create_sheet("Validacion")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Incidencias de Validacion"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    header_row = 3
    headers = ["Archivo", "Fila", "Nivel", "Mensaje"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _style_header_row(ws, header_row, len(headers))

    start = header_row + 1
    level_fill = {"error": PatternFill("solid", fgColor=RED_BG), "warning": PatternFill("solid", fgColor=YELLOW_BG)}
    level_font = {"error": Font(color=RED_TEXT), "warning": Font(color=YELLOW_TEXT)}

    for i, issue in enumerate(issues):
        r = start + i
        ws.cell(row=r, column=1, value=issue.archivo)
        ws.cell(row=r, column=2, value=issue.fila)
        ws.cell(row=r, column=3, value=issue.nivel)
        ws.cell(row=r, column=4, value=issue.mensaje)
        fill = level_fill.get(issue.nivel)
        font = level_font.get(issue.nivel)
        for col_idx in range(1, 5):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if font:
                cell.font = font

    if not issues:
        ws.cell(row=start, column=1, value="Sin incidencias — todos los movimientos son validos.")
        ws.cell(row=start, column=1).font = Font(color=GREEN_TEXT, bold=True)
        ws.cell(row=start, column=1).fill = PatternFill("solid", fgColor=GREEN_BG)

    _autosize(ws, {"A": 26, "B": 10, "C": 12, "D": 70})
    return ws


def build_report(movements: list[Movement], issues: list[ParseIssue]) -> bytes:
    """Builds the full workbook in memory and returns its bytes."""
    wb = Workbook()
    _write_movements_sheet(wb, movements)
    _write_resumen_sheet(wb, len(movements))
    _write_consulta_sheet(wb)
    _write_validacion_sheet(wb, issues)

    wb.active = 0  # land on Movimientos when opened
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
