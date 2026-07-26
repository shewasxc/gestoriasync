"""Generates the demo_files/ pack used to try out the app: one BBVA PDF
statement (realistic layout), one Santander XLSX, one CaixaBank CSV, and
one generic invoice ledger (to exercise the fuzzy-column fallback parser).
Real Spanish expense concepts (Alquiler, Mercadona, Gasolina, Impuestos...),
DD/MM/YYYY dates, comma-decimal amounts, and a deliberate mix of 3 valid +
2 invalid NIF/CIF spread across the files so the validation highlighting
has something to catch in every file.

Run: venv/Scripts/python.exe demo_files/generate_demo_data.py
"""
from __future__ import annotations

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable

HERE = os.path.dirname(os.path.abspath(__file__))

_DNI_LETTERS = "TRWAGMYFPDXBNJZSQVHLCKE"


def valid_dni(number: str) -> str:
    return number + _DNI_LETTERS[int(number) % 23]


def valid_cif(letter: str, number: str) -> str:
    even_sum = sum(int(number[i]) for i in range(1, 7, 2))
    odd_sum = 0
    for i in range(0, 7, 2):
        doubled = int(number[i]) * 2
        odd_sum += doubled // 10 + doubled % 10
    unit = (even_sum + odd_sum) % 10
    numeric_control = str((10 - unit) % 10)
    letter_control = "JABCDEFGHI"[(10 - unit) % 10]
    control = letter_control if letter in "KPQS" else numeric_control
    return f"{letter}{number}{control}"


# --- 3 valid + 2 invalid NIF/CIF, reused across the three files -----------
V_DNI = valid_dni("30456789")          # persona fisica (autonomo) - valida
V_CIF_SL = valid_cif("B", "8172645")   # sociedad limitada - valida
V_CIF_SA = valid_cif("A", "4809215")   # sociedad anonima - valida

I_DNI_BAD = "30456789X"                # letra de control incorrecta a proposito
I_CIF_BAD = "B81726401"                # digito de control erroneo a proposito (el correcto seria 8)

print("Valid  DNI:", V_DNI)
print("Valid  CIF (SL):", V_CIF_SL)
print("Valid  CIF (SA):", V_CIF_SA)
print("Invalid DNI:", I_DNI_BAD)
print("Invalid CIF:", I_CIF_BAD)


# ---------------------------------------------------------------------------
# 1) BBVA — PDF con maquetacion realista de extracto bancario
# ---------------------------------------------------------------------------
def build_bbva_pdf() -> None:
    path = os.path.join(HERE, "extracto_BBVA_demo.pdf")
    doc = SimpleDocTemplate(
        path, pagesize=A4,
        topMargin=1.6 * cm, bottomMargin=1.6 * cm, leftMargin=1.8 * cm, rightMargin=1.8 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("BBVATitle", parent=styles["Title"], textColor=colors.HexColor("#004481"), fontSize=20)
    sub_style = ParagraphStyle("BBVASub", parent=styles["Normal"], textColor=colors.HexColor("#404040"), fontSize=9.5)

    elements = [
        Paragraph("BBVA", title_style),
        Paragraph("Extracto integral de cuenta corriente", sub_style),
        Spacer(1, 0.15 * cm),
        HRFlowable(width="100%", thickness=1.2, color=colors.HexColor("#004481")),
        Spacer(1, 0.35 * cm),
        Paragraph("<b>Titular:</b> Estudio Creativo Levante SL &nbsp;&nbsp;&nbsp; <b>IBAN:</b> ES91 0182 4324 4102 0800 1234", sub_style),
        Paragraph("<b>Periodo:</b> 01/01/2026 - 30/06/2026 &nbsp;&nbsp;&nbsp; <b>Oficina:</b> 4324 - Valencia Centro", sub_style),
        Spacer(1, 0.6 * cm),
    ]

    header = ["Fecha", "Concepto", "Importe", "NIF/CIF", "Saldo"]
    data = [header] + [
        ["05/01/2026", "ALQUILER LOCAL COMERCIAL ENE", "-1.250,50 €", V_CIF_SL, "22.340,10 €"],
        ["08/01/2026", "MERCADONA S.A. COMPRA OFICINA", "-63,45 €", V_CIF_SA, "22.276,65 €"],
        ["14/01/2026", "GASOLINA REPSOL AVDA FRANCIA", "-58,20 €", V_DNI, "22.218,45 €"],
        ["20/01/2026", "AGENCIA TRIBUTARIA MODELO 303", "-940,15 €", V_CIF_SL, "21.278,30 €"],
        ["03/02/2026", "TRANSFERENCIA RECIBIDA CLIENTE", "3.200,00 €", V_CIF_SA, "24.478,30 €"],
        ["05/02/2026", "ALQUILER LOCAL COMERCIAL FEB", "-1.250,50 €", I_DNI_BAD, "23.227,80 €"],
        ["11/02/2026", "MERCADONA S.A. COMPRA OFICINA", "-71,10 €", V_CIF_SA, "23.156,70 €"],
        ["18/02/2026", "GASOLINA CEPSA RONDA NORD", "-46,90 €", V_DNI, "23.109,80 €"],
        ["10/03/2026", "IMPUESTOS AYUNTAMIENTO IBI LOCAL", "-410,00 €", V_CIF_SL, "22.699,80 €"],
        ["22/03/2026", "NOMINA MARZO EQUIPO", "-2.480,00 €", V_CIF_SL, "20.219,80 €"],
    ]

    table = Table(data, colWidths=[2.3 * cm, 6.6 * cm, 2.9 * cm, 3.0 * cm, 2.9 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#004481")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.3),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B9C4CE")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF3F8")]),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph("Documento generado con fines de demostracion. BBVA no es responsable de su contenido.", sub_style))
    doc.build(elements)


# ---------------------------------------------------------------------------
# 2) Santander — XLSX con cabecera de banco y formato de columnas propio
# ---------------------------------------------------------------------------
def build_santander_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Banco Santander — Extracto de Movimientos"
    ws["A1"].font = Font(bold=True, size=14, color="EC0000")
    ws.merge_cells("A2:F2")
    ws["A2"] = "Cuenta: ES76 0049 1234 5610 0012 3456   |   Trimestre 1-2 2026"
    ws["A2"].font = Font(size=9, color="595959")

    header_row = 4
    headers = ["Fecha Operacion", "Fecha Valor", "Concepto", "Importe EUR", "NIF/CIF", "Saldo"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="EC0000")
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("07/01/2026", "07/01/2026", "ALQUILER NAVE ALMACEN", -980.00, V_CIF_SA, 15230.00),
        ("12/01/2026", "13/01/2026", "MERCADONA COMPRA SUMINISTROS", -54.85, V_DNI, 15175.15),
        ("19/01/2026", "19/01/2026", "GASOLINERA REPSOL N-332", -62.30, V_CIF_SL, 15112.85),
        ("28/02/2026", "28/02/2026", "IMPUESTOS MODELO 130 IRPF", -615.40, V_CIF_SA, 14497.45),
        ("05/03/2026", "06/03/2026", "TRANSFERENCIA CLIENTE FACTURA 0032", 2750.00, I_CIF_BAD, 17247.45),
        ("14/03/2026", "14/03/2026", "ALQUILER NAVE ALMACEN", -980.00, V_CIF_SA, 16267.45),
        ("22/04/2026", "22/04/2026", "MERCADONA COMPRA SUMINISTROS", -48.20, V_DNI, 16219.25),
        ("30/04/2026", "30/04/2026", "GASOLINA CEPSA AUTOVIA A-7", -71.55, V_CIF_SL, 16147.70),
        ("18/05/2026", "18/05/2026", "SEGURO MAPFRE AUTONOMOS", -95.00, V_CIF_SA, 16052.70),
        ("30/06/2026", "30/06/2026", "NOMINA JUNIO EQUIPO", -2480.00, V_CIF_SL, 13572.70),
    ]
    for i, r in enumerate(rows):
        row_idx = header_row + 1 + i
        for col, value in enumerate(r, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col == 4:
                cell.number_format = '#,##0.00" €"'

    for col_letter, width in zip("ABCDEF", (14, 14, 34, 14, 14, 14)):
        ws.column_dimensions[col_letter].width = width

    wb.save(os.path.join(HERE, "movimientos_Santander_demo.xlsx"))


# ---------------------------------------------------------------------------
# 3) CaixaBank — CSV con cabecera 'Saldo disponible' (firma del banco)
# ---------------------------------------------------------------------------
def build_caixabank_csv() -> None:
    rows = [
        ["04/01/2026", "ALQUILER OFICINA COWORKING", "-620,00", V_CIF_SL, "9.840,20"],
        ["09/01/2026", "MERCADONA COMPRA MATERIAL", "-39,70", V_CIF_SA, "9.800,50"],
        ["16/01/2026", "GASOLINA REPSOL RONDA LITORAL", "-55,40", V_DNI, "9.745,10"],
        ["27/02/2026", "IMPUESTOS AGENCIA TRIBUTARIA IVA", "-730,25", V_CIF_SL, "9.014,85"],
        ["11/03/2026", "INGRESO VENTA SERVICIOS CLIENTE", "1.980,00", V_CIF_SA, "10.994,85"],
        ["23/03/2026", "MERCADONA COMPRA MATERIAL", "-42,15", I_CIF_BAD, "10.952,70"],
        ["05/04/2026", "ALQUILER OFICINA COWORKING", "-620,00", V_CIF_SL, "10.332,70"],
        ["19/05/2026", "GASOLINA CEPSA B-23", "-49,80", V_DNI, "10.282,90"],
        ["02/06/2026", "IMPUESTOS MODELO 111 RETENCIONES", "-215,60", V_CIF_SL, "10.067,30"],
        ["28/06/2026", "SEGURIDAD SOCIAL TGSS TRIMESTRE", "-410,00", V_CIF_SA, "9.657,30"],
    ]
    df = pd.DataFrame(rows, columns=["Fecha", "Concepto", "Importe", "NIF/CIF", "Saldo disponible"])
    df.to_csv(os.path.join(HERE, "movimientos_CaixaBank_demo.csv"), index=False, sep=",", encoding="utf-8")


# ---------------------------------------------------------------------------
# 4) Generic ledger — XLSX with non-bank column names, to exercise the
#    fuzzy-matching fallback parser rather than any bank signature.
# ---------------------------------------------------------------------------
def build_generic_invoices_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.append(["Fecha Factura", "Descripcion", "Total Factura", "Identificador Fiscal Cliente"])
    rows = [
        ("15/01/2026", "Factura proveedor material informatico", -320.00, V_CIF_SL),
        ("28/02/2026", "Factura cliente servicios de consultoria", 1850.00, V_DNI),
        ("10/03/2026", "Factura proveedor suministro electrico", -410.75, V_CIF_SA),
        ("22/04/2026", "Factura cliente desarrollo software", 2600.00, V_DNI),
        ("05/05/2026", "Factura proveedor transporte mercancias", -145.30, I_DNI_BAD),
        ("19/06/2026", "Factura cliente mantenimiento anual", 990.00, V_CIF_SL),
        ("07/07/2026", "Factura proveedor alquiler almacen", -700.00, I_CIF_BAD),
    ]
    for r in rows:
        ws.append(list(r))
    wb.save(os.path.join(HERE, "facturas_genericas_demo.xlsx"))


def main() -> None:
    build_bbva_pdf()
    build_santander_xlsx()
    build_caixabank_csv()
    build_generic_invoices_xlsx()
    print("\nArchivos generados en", HERE)


if __name__ == "__main__":
    main()
